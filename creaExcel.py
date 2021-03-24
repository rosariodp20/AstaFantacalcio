import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def CreaTabella(squadre):   #mi importo tutte le istanze delle squadre

    TabellaFinale=openpyxl.Workbook()
    Foglio=TabellaFinale.active
    Foglio.title="Tabella Squadre"

    numSquadre=len(squadre)

    c=1
    r=3

    for squadra in squadre:

        listaPortieri=squadra.getListaPortieri()
        listaDifensori=squadra.getListaDifensori()
        listaCentrocampisti=squadra.getListaCentrocampisti()
        listaAttaccanti=squadra.getListaAttaccanti()

        for portiere in listaPortieri:

            nomePort=portiere.getNome()
            squadraPort=portiere.getSquadra()
            valorePort=portiere.getValore()

            Foglio.cell(row=r,column=c).value=nomePort
            Foglio.cell(row=r,column=c+1).value=squadraPort
            Foglio.cell(row=r,column=c+2).value=valorePort

            r=r+1

        r=7
        for difensore in listaDifensori:

            nomeDif=difensore.getNome()
            squadraDif=difensore.getSquadra()
            valoreDif=difensore.getValore()

            Foglio.cell(row=r,column=c).value=nomeDif
            Foglio.cell(row=r,column=c+1).value=squadraDif
            Foglio.cell(row=r,column=c+2).value=valoreDif

            r=r+1

        r=16
        for centrocampista in listaCentrocampisti:

            nomeCentr=centrocampista.getNome()
            squadraCentr=centrocampista.getSquadra()
            valoreCentr=centrocampista.getValore()

            Foglio.cell(row=r,column=c).value=nomeCentr
            Foglio.cell(row=r,column=c+1).value=squadraCentr
            Foglio.cell(row=r,column=c+2).value=valoreCentr

            r=r+1

        r=25
        for attaccante in listaAttaccanti:

            nomeAtt=attaccante.getNome()
            squadraAtt=attaccante.getSquadra()
            valoreAtt=attaccante.getValore()

            Foglio.cell(row=r,column=c).value=nomeAtt
            Foglio.cell(row=r,column=c+1).value=squadraAtt
            Foglio.cell(row=r,column=c+2).value=valoreAtt

            r=r+1

        r=3
        c=c+3


    a=1
    b=1
    c=1
    d=3

    for i in range(numSquadre):
        Foglio.merge_cells(start_row=a,start_column=b,end_row=c,end_column=d)
        Foglio.merge_cells(start_row=a+1,start_column=b,end_row=c+1,end_column=d)
        Foglio.merge_cells(start_row=a+5,start_column=b,end_row=c+5,end_column=d)
        Foglio.merge_cells(start_row=a+14,start_column=b,end_row=c+14,end_column=d)
        Foglio.merge_cells(start_row=a+23,start_column=b,end_row=c+23,end_column=d)
        Foglio.cell(row=2,column=b).value="Portieri"
        Foglio.cell(row=2,column=b).font=Font(bold=True,size=12,color='FFFF6600')
        Foglio.cell(row=2,column=b).alignment=Alignment(horizontal='center')
        Foglio.cell(row=6,column=b).value="Difensori"
        Foglio.cell(row=6,column=b).font=Font(bold=True,size=12,color='FF0000FF')
        Foglio.cell(row=6,column=b).alignment=Alignment(horizontal='center')
        Foglio.cell(row=15,column=b).value="Centrocampisti"
        Foglio.cell(row=15,column=b).font=Font(bold=True,size=12,color='FF008000')
        Foglio.cell(row=15,column=b).alignment=Alignment(horizontal='center')
        Foglio.cell(row=24,column=b).value="Attaccanti"
        Foglio.cell(row=24,column=b).font=Font(bold=True,size=12,color='FFFF0000')
        Foglio.cell(row=24,column=b).alignment=Alignment(horizontal='center')

        letteraColonna=get_column_letter

        SquadraTemp=squadre[i]
        nomeSquadra=SquadraTemp.getNomeSquadra()
        Foglio.cell(row=1,column=b).value=nomeSquadra
        Foglio.cell(row=1,column=b).font=Font(bold=True,size=18)
        Foglio.cell(row=1,column=b).alignment=Alignment(horizontal='center')

        b=b+3
        d=d+3


    for col in Foglio.columns:
        column=col[0].column
        if column in [1,4,7,10,13,16,19,22,25,28,31,34,37]:
            let=get_column_letter(column)
            Foglio.column_dimensions[let].width="17"

    TabellaFinale.save('SquadreFantacalcio.xlsx')
