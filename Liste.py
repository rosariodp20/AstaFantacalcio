from openpyxl import load_workbook


workbook=load_workbook(filename='Quotazioni_Fantacalcio.xlsx')     #Apro il file excel in lettura
sheet=workbook['Tutti']                                            #Mi posiziono sul foglio "Tutti" del file excel che ho aperto

#sheet[C2].value      #Modo per selezionare il contenuto di un cella Excel
#sheet.cell(row=5, column=2).value          #Modo alternativo per selezionare il contenuto di una cella Excel

Lista_Att=list()
Lista_Centr=list()
Lista_Dif=list()
Lista_Port=list()

i=3
while i<600:
    B='B'+str(i)
    C='C'+str(i)
    if sheet[B].value=='A':
        Lista_Att.append(sheet[C].value)
    elif sheet[B].value=='C':
        Lista_Centr.append(sheet[C].value)
    elif sheet[B].value=='P':
        Lista_Port.append(sheet[C].value)
    elif sheet[B].value=='D':
        Lista_Dif.append(sheet[C].value)

    i=i+1

print(Lista_Att)
print(Lista_Centr)
print(Lista_Dif)
print(Lista_Port)
