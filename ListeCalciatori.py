from openpyxl import load_workbook
from Calciatore import Calciatore
import random



def CreaIstanzeCalciatori():

    global ListaAtt, ListaCentr, ListaDif, ListaPort

    workbook=load_workbook(filename='Quotazioni_Fantacalcio.xlsx')     #Apro il file excel in lettura
    sheet=workbook['Tutti']                                            #Mi posiziono sul foglio "Tutti" del file excel che ho aperto

    #sheet[C2].value      #Modo per selezionare il contenuto di un cella Excel
    #sheet.cell(row=5, column=2).value          #Modo alternativo per selezionare il contenuto di una cella Excel

    ListaAtt=list()
    ListaCentr=list()
    ListaDif=list()
    ListaPort=list()

    i=3
    while i<600:
        B='B'+str(i)
        C='C'+str(i)
        D='D'+str(i)

        if sheet[B].value=='A':
            Attaccante=Calciatore(sheet[C].value, sheet[D].value, 'A')
            ListaAtt.append(Attaccante)

        elif sheet[B].value=='C':
            Centrocampista=Calciatore(sheet[C].value, sheet[D].value, 'C')
            ListaCentr.append(Centrocampista)

        elif sheet[B].value=='P':
            Portiere=Calciatore(sheet[C].value, sheet[D].value, 'P')
            ListaPort.append(Portiere)

        elif sheet[B].value=='D':
            Difensore=Calciatore(sheet[C].value, sheet[D].value, 'D')
            ListaDif.append(Difensore)

        i=i+1

def getAttaccanti():
    random.shuffle(ListaAtt)
    return ListaAtt

def getCentrocampisti():
    random.shuffle(ListaCentr)
    return ListaCentr

def getDifensori():
    random.shuffle(ListaDif)
    return ListaDif

def getPortieri():
    random.shuffle(ListaPort)
    return ListaPort
